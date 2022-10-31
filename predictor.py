
from openpyxl import load_workbook
import sys
import numpy as np
from scipy.optimize import minimize
import statistics

"""
Input:
     - filename: str of an excel file
Output data:
    - mkt_prices: [100, 200, ..,] len(marker_prices) = num_months
    - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
    - analysts: set {analyst_code, analyst_code, ...}
"""
def GetData(filename):
    workbook = load_workbook(filename, data_only=True)
    ws = workbook["Sheet1"]

    source_data = ws['B6':'BK200']
    mkt_price_data = ws['B3':'BK3'][0]

    analyst_col = 0
    monthly_data = []
    mkt_prices = []
    analysts = set()

    while analyst_col < len(source_data[0])-1:
        current_month_data = dict()
        for row in range(len(source_data)):
            cell_value = source_data[row][analyst_col].value
            if cell_value != 0 and not cell_value is None:
                current_month_data[cell_value] = source_data[row][analyst_col+1].value
                analysts.add(cell_value)
        monthly_data.append(current_month_data)
        mkt_prices.append(mkt_price_data[analyst_col+1].value)
        analyst_col+=2

    # print("monthly data")
    # print(monthly_data)
    # print("mkt_prices")
    # print(mkt_prices)
    # print("analysts")
    # print(analysts)
    return monthly_data, mkt_prices, analysts

"""
Input:
    - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
    - analysts: set {analyst_code, analyst_code, ...}
Output data:
    - analyst_data: [[100,200,300],[150,150,400],[...],] len(analyst_data) = num_analysts, len(analyst_data[0]) = num_months
"""
def Clean(monthly_data, analysts):
    def CleanSingleAnalyst(single_analyst_monthly_prections):
        for idx in range(len(single_analyst_monthly_prections)):
            price_target = single_analyst_monthly_prections[idx]
            if not price_target is None:
                continue

            def GetNextVal(idx):
                next_idx=idx+1
                while next_idx < len(single_analyst_monthly_prections):
                    val = single_analyst_monthly_prections[next_idx]
                    if not val is None:
                        return val, next_idx-idx+1
                    next_idx+=1
                return None, None

            next_val, next_distance = GetNextVal(idx)

            #no previous, yes next
            if idx==0 and not next_val is None:
                single_analyst_monthly_prections[idx] = next_val
            #yes previous, yes next
            elif idx != 0 and not next_val is None:
                prev_val = single_analyst_monthly_prections[idx-1]
                single_analyst_monthly_prections[idx] = prev_val + ((next_val-prev_val)/next_distance)
            #yes previous, no next
            elif idx != 0 and next_val is None:
                prev_val=single_analyst_monthly_prections[idx-1]
                single_analyst_monthly_prections[idx] = prev_val
        return single_analyst_monthly_prections

    sorted_analysts = sorted(list(analysts))
    cleaned_data = []
    for analyst in sorted_analysts:
        single_analyst_monthly_prections = []
        contiuous_none = 0
        previous_none = False
        for curr_month in monthly_data:
            if not analyst in curr_month or curr_month[analyst] is None or curr_month[analyst] == '@NA':
                single_analyst_monthly_prections.append(None)
                if previous_none:
                    contiuous_none+=1
                    if contiuous_none > 5:
                        break
                else:
                    contiuous_none=1
                previous_none=True
            else:
                single_analyst_monthly_prections.append(curr_month[analyst])
                previous_none = False
        if contiuous_none <= 5:
            single_analyst_monthly_prections = CleanSingleAnalyst(single_analyst_monthly_prections) if contiuous_none > 0 else single_analyst_monthly_prections
            cleaned_data.append(single_analyst_monthly_prections)
    return cleaned_data

"""
Input:
    - analyst_data: [[100,200,300],[150,150,400],[...],] len(analyst_data) = num_analysts, len(analyst_data[0]) = num_months
    - mkt_prices: [100, 200, ..,] len(marker_prices) = num_months
Output:
    - learned_weights: [float, ..] len = num_analysts
    - average_month, standard_deviation: floats
"""
def LearnWeights(analyst_data, mkt_prices):
    num_analysts = len(analyst_data)
    num_months = len(analyst_data[0])
    analyst_weights = [1/num_analysts for i in range(num_analysts)] # if 10 analysts, [0.1 0.1 0.1 0.1 ..]
    analyst_weights = np.array(analyst_weights)
    analyst_data = np.array(analyst_data)

    def scoring_function(analyst_weights, give_stdv_and_avg=False):
        predicted_prices = np.matmul(analyst_weights, analyst_data)
        time_to_reach = list()
        count24 = 0

        for idx in range(num_months):
            found = False
            mkt_price_idx = idx+1
            while mkt_price_idx < len(mkt_prices):
                if (predicted_prices[idx] >= mkt_prices[idx] and int(mkt_prices[mkt_price_idx]) >= int(predicted_prices[idx])) or (predicted_prices[idx] < mkt_prices[idx] and int(mkt_prices[mkt_price_idx]) <= int(predicted_prices[idx])):
                    found=True
                    time_to_reach.append((mkt_price_idx - idx) - (abs(mkt_prices[mkt_price_idx] - predicted_prices[idx])/abs(mkt_prices[mkt_price_idx] - mkt_prices[mkt_price_idx-1]) ))
                    break
                mkt_price_idx+=1
            if not found:
                time_to_reach.append(24)
                count24+=1

        stdv = statistics.stdev(time_to_reach)
        av = statistics.mean(time_to_reach)

        score = av + 1.3*stdv + 4*count24

        if give_stdv_and_avg:
            print("average:", av)
            print("stdev:", stdv)

            return score, av, stdv

        return score

    print("before learning")
    _, _, _ =  scoring_function(analyst_weights, True)

    result = minimize(scoring_function, analyst_weights, method='SLSQP', options={'disp': True})

    print("After learning:")
    score, average_months, stdv = scoring_function(result.x, True)

    return result.x, average_months, stdv

"""
Input:
    - leanrned_weights: [float, ..] len = num_analysts
    - target_prices: [float, ...] len = num_analysts
Output:
    - predicted price floats
"""
def Predict():
    pass

filename = sys.argv[1]
analyst_data, mkt_prices, analysts = GetData(filename)

cleaned_data = Clean(analyst_data, analysts)

training_data = []
for row in cleaned_data:
    training_data.append(row[:-6])

leanrned_weights, average_months, stdv = LearnWeights(training_data, mkt_prices)

print(leanrned_weights)
#
# predicted_price = Predict(leanrned_weights, analyst_data[-1])
#
# print("predicted price", predicted_price)
# print("estimated months to reach", average_months)

# """
# """
# Input:
#      - filename: str of an excel file
# Output data:
#     - market_prices: [100, 200, ..,] len(marker_prices) = num_months
#     - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
#     - analysts: set {analyst_code, analyst_code, ...}
# """
# def GetData(filename):

#     workbook = load_workbook(filename, data_only=True)
 
#     ws = workbook["Sheet1"]
#     source_data = ws['B6' : 'BK200']
#     market_price_data = ws['B3' : 'BK3'][0]
#     analyst_col = 0
#     monthly_data = []
#     market_prices = []
#     analysts = set()
#     while analyst_col < len(source_data[0]) - 1:
#         current_month_data = dict()
#         for row in range(len(source_data)):
#             cell_value = source_data[row][analyst_col].value
#             if cell_value != 0 and cell_value is not None:
#                 current_month_data[cell_value] = source_data[row][analyst_col + 1].value
#                 analysts.add(cell_value)
#         monthly_data.append(current_month_data)
#         market_prices.append(market_price_data[analyst_col + 1].value)
#         #Goes to the next analysts
#         analyst_col += 2
#     # print("Monthly data")
#     # print(monthly_data)
#     # print("Market prices")
#     # print(market_prices)
#     # print("analysts")
#     # print(analysts)
#     return monthly_data,market_prices,analysts

# # """
# # Input:
# #     - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
# #     - analysts: set {analyst_code, analyst_code, ...}
# # Output data:
# #     - analyst_data: [[100,200,300],[150,150,400],[...],] len(analyst_data) = num_analysts, len(analyst_data[0]) = num_months
# # """
# def Clean(monthly_data, analysts):
#     def Cleaned_single_analyst(single_analyst_monthly_prediction):
#         for index in range(len(single_analyst_monthly_prediction)):
#             price_target = single_analyst_monthly_prediction[index]
#             if not price_target is None:
#                 continue
#             def GetNextVal(index):
#                 next_index = index + 1
#                 while next_index < len(single_analyst_monthly_prediction):
#                     val = single_analyst_monthly_prediction[next_index]
#                     if not val is None:
#                         return val, next_index - index +1
#                     next_index += 1
#                 return None, None
#             #3 scenarios
#             #no precvious, yes next 
#             next_val, next_distance = GetNextVal(index)
#             if index == 0 and not next_val is None:
#                 single_analyst_monthly_prediction[index] = next_val
#             elif  index !=0 and not next_val is None:
#                 prev_value = single_analyst_monthly_prediction[index - 1]
#                 single_analyst_monthly_prediction[index] = prev_value + ((next_val - prev_value) / next_distance)
#             elif index != 0 and next_val is None:
#                 prev_value = single_analyst_monthly_prediction[index - 1]
#                 single_analyst_monthly_prediction[index] = prev_value
#         return single_analyst_monthly_prediction
#             #yes previous, no next
#             #yes previous, yes next
#     sorted_analysts = sorted(list(analysts))
#     cleaned_data = []
    
#     for analyst in sorted_analysts:
#         single_analyst_monthly_prediction = []
#         nones_count = 0
#         previous_none = False
#         for curr_month in monthly_data:
#             if not analyst in curr_month or curr_month[analyst] is None or curr_month[analyst] == "@NA":
#                 single_analyst_monthly_prediction.append(None) 
#                 if previous_none:
#                     nones_count += 1
#                     if nones_count > 5:
#                         break
#                 else:
#                     nones_count = 1
#                 previous_none = True

#                 # nones_count += 1
#             else:
#                 single_analyst_monthly_prediction.append(curr_month[analyst])
#                 previous_none = False
#         if nones_count <= 5: 
#             single_analyst_monthly_prediction = Cleaned_single_analyst(single_analyst_monthly_prediction) if nones_count > 0 else single_analyst_monthly_prediction
#             cleaned_data.append(single_analyst_monthly_prediction)
#     return cleaned_data


        
#     pass

# # Input:
# #     - monthly_data: [{analyst_code: target_price, ...},{...}, ...] len(monthly_data) = num_months,
# #     - analysts: set {analyst_code, analyst_code, ...}
# # Output data:
# #     - analyst_data: [[100,200,300],[150,150,400],[...],] len(analyst_data) = num_analysts, len(analyst_data[0]) = num_months
# # """
# #Learn the weights on which analysts to include 
# def LearnWeights(analyst_data, market_prices):
#     num_analysts = len(analyst_data)
#     num_months = len(analyst_data[0])
#     analyst_weights = [1/num_analysts for i in range(num_analysts)]
#     analyst_weights = np.array(analyst_weights)
#     analyst_data = np.array(analyst_data)
    
#     def scoring_function(analyst_weights, give_std_and_mean = False):
#         predicted_prices = np.matmul(analyst_weights, analyst_data)
#         time_to_reach_prices = list()
#         count24 = 0
#         for index in range(num_months):
#             found = False
#             mkt_prices_index = index +1
#             while mkt_prices_index < len(mkt_prices):
#                 if (predicted_prices[index] >= market_prices[index] and int(market_prices[mkt_prices_index]) >= int(predicted_prices[index])) or (predicted_prices[index] < market_prices[index] and int(market_prices[mkt_prices_index]) <= int(predicted_prices[index])):
#                     time_to_reach_prices.append((mkt_prices_index - index) - (abs(market_prices[mkt_prices_index] - predicted_prices[index])/abs(market_prices[mkt_prices_index] - market_prices[mkt_prices_index-1]) ))
#                     found = True
#                     break
#                 mkt_prices_index += 1
#             if not found:
#                 time_to_reach_prices.append(24)
#                 count24 += 1
#             stdev = statistics.stdev(time_to_reach_prices)
#             mean = statistics.mean(time_to_reach_prices)
#             score = mean + 1.3 * stdev + count24*4
#             if give_std_and_mean:
#                     print("average:", mean)
#                     print("stdev:", stdev)
#                     return score, mean, stdev
#             # return score
#         print("Before learning")
#         x = scoring_function(analyst_weights, give_std_and_mean = True)

#         result = minimize(x, analyst_weights, method='SLSQP', options={'disp': True})
#         print("After learning")
#         score, mean, stdev = scoring_function(result.x, give_std_and_mean = True)

#         return result.x, score, mean, stdev



#     # Input:
# #     - leanrned_weights: [float, ..] len = num_analysts
# #     - target_prices: [float, ...] len = num_analysts
# # Output:
# #     - predicted price floats
# # """
    

# def Predict():
#     pass

# filename = sys.argv[1]
# analyst_data, mkt_prices, analysts = GetData(filename)
# clean_data = Clean(analyst_data, analysts)
# training_data = []
# for row in clean_data:
#     #ignores last 6 months
#     training_data.append(row[:-6])
# print(mkt_prices)

# learned_weights, avg_months, stdev = LearnWeights(training_data, mkt_prices)
# print(learned_weights)




